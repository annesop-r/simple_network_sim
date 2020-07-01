from pathlib import Path
from urllib.request import urlretrieve
import openpyxl


# This script is simultaneously poorly written and overengineered.
# I expected to have to do some complicated interpolation of school locations into
# health boards before I saw that health boards are a super set of local authorities ant
# that local authorities are not scplit across health boards - thus the overengineering.
# The poorly written part has no excuse.....

# NOTE:
# We need the following spreadsheet because it contains data on each school's capacity in each local authority
# https://www2.gov.scot/Resource/0054/00548378.xlsx

# Other useful data on school level statistics
# https://www.gov.scot/binaries/content/documents/govscot/publications/statistics/2020/03/school-level-summary-statistics-2019/documents/school-level-summary-statistics-2019/school-level-summary-statistics-2019/govscot%3Adocument/School%2Blevel%2Bsummary%2Bstatistics%2B2019%2Bupdated%2BMay%2B2020.xlsx


healthBoardId = {"S08000015":"NHS Ayrshire and Arran",
"S08000016":"NHS Borders",
"S08000017":"NHS Dumfries and Galloway",
"S08000019":"NHS Forth Valley",
"S08000020":"NHS Grampian",
"S08000022":"NHS Highland",
"S08000024":"NHS Lothian",
"S08000025":"NHS Orkney",
"S08000026":"NHS Shetland",
"S08000028":"NHS Western Isles",
"S08000029":"NHS Fife",
"S08000031":"NHS Greater Glasgow and Clyde",
"S08000032":"NHS Lanarkshire",
"S08000030":"NHS Tayside"}



# Link between Health Board and Local Authorities  (https://en.wikipedia.org/wiki/Subdivisions_of_Scotland#Health)
healthBoardsAndLocalAuthoriesMap = {"Ayrshire and Arran" : ["East Ayrshire","North Ayrshire","South Ayrshire"],
"Borders" : ["Scottish Borders"],
"Dumfries and Galloway" : ["Dumfries and Galloway"],
"Fife" : ["Fife"],
"Forth Valley" : ["Clackmannanshire", "Falkirk", "Stirling"],
"Grampian" : ["Aberdeenshire", "City of Aberdeen", "Moray"],
"Greater Glasgow and Clyde" : ["City of Glasgow", "East Dunbartonshire", "East Renfrewshire", "Inverclyde", "Renfrewshire", "West Dunbartonshire"],
"Highland" : ["Argyll and Bute", "Highland"],
"Lanarkshire" : ["North Lanarkshire", "South Lanarkshire"],
"Lothian" : ["City of Edinburgh", "East Lothian", "Midlothian", "West Lothian"],
"Orkney" : ["Orkney Islands"],
"Shetland" : ["Shetland Islands"],
"Tayside" : ["Angus", "City of Dundee", "Perth and Kinross"],
"Western Isles" : ["Western Isles (Na h-Eileanan Siar)"]}



class LocalAuthority:
  def __init__(self, name, population, area, density):
    self.name = name
    self.population = population
    self.area = area
    self.density = density
    self.primary_schools = 0
    self.secondary_schools = 0
    self.total_schools = 0
    self.primary_school_total = 0
    self.secondary_school_total = 0
    self.special_school_total = 0
    self.primary_school_pupils = []
    self.secondary_school_pupils = []
    self.special_school_pupils = []


# Take the table from widipedia (https://en.wikipedia.org/wiki/Subdivisions_of_Scotland#Council_areas)
# copy it to a file called tmp and manipulate is as 
# sed -e's^,^^g' tmp | awk 'FS="\t" {print "\""$2"\" : LocalAuthority(\""$2"\","$3","$5","$6")\,"}'
# with a little manipulation for the name of Western Isles, City of Glasgow 

localAuthorities = {
"City of Glasgow" : LocalAuthority("City of Glasgow",626410,174.7,3586),
"City of Edinburgh" : LocalAuthority("City of Edinburgh",518500,263.4,1969),
"Fife" : LocalAuthority("Fife",371910,1325,281),
"North Lanarkshire" : LocalAuthority("North Lanarkshire",340180,469.9,724),
"South Lanarkshire" : LocalAuthority("South Lanarkshire",319020,1772,180),
"Aberdeenshire" : LocalAuthority("Aberdeenshire",261470,6313,41),
"Highland" : LocalAuthority("Highland",235540,25657,9),
"City of Aberdeen" : LocalAuthority("City of Aberdeen",227560,185.7,1225),
"West Lothian" : LocalAuthority("West Lothian",182140,427.7,426),
"Renfrewshire" : LocalAuthority("Renfrewshire",177790,261.5,680),
"Falkirk" : LocalAuthority("Falkirk",160340,297.4,539),
"Perth and Kinross" : LocalAuthority("Perth and Kinross",151290,5286,29),
"Dumfries and Galloway" : LocalAuthority("Dumfries and Galloway",148790,6427,23),
"City of Dundee" : LocalAuthority("City of Dundee",148750,59.83,2486),
"North Ayrshire" : LocalAuthority("North Ayrshire",135280,885.4,153),
"East Ayrshire" : LocalAuthority("East Ayrshire",121840,1262,97),
"Angus" : LocalAuthority("Angus",116040,2182,53),
"Scottish Borders" : LocalAuthority("Scottish Borders",115270,4732,24),
"South Ayrshire" : LocalAuthority("South Ayrshire",112550,1222,92),
"East Dunbartonshire" : LocalAuthority("East Dunbartonshire",108330,174.5,621),
"East Lothian" : LocalAuthority("East Lothian",105790,679.2,156),
"Moray" : LocalAuthority("Moray",95520,2238,43),
"East Renfrewshire" : LocalAuthority("East Renfrewshire",95170,174.2,546),
"Stirling" : LocalAuthority("Stirling",94330,2187,43),
"Midlothian" : LocalAuthority("Midlothian",91340,353.7,258),
"West Dunbartonshire" : LocalAuthority("West Dunbartonshire",89130,158.8,561),
"Argyll and Bute" : LocalAuthority("Argyll and Bute",86260,6909,12),
"Inverclyde" : LocalAuthority("Inverclyde",78150,160.5,487),
"Clackmannanshire" : LocalAuthority("Clackmannanshire",51400,159.0,323),
"Western Isles (Na h-Eileanan Siar)" : LocalAuthority("Na h-Eileanan Siar (Western Isles)",26830,3059,9),
"Shetland Islands" : LocalAuthority("Shetland Islands",22990,1468,16),
"Orkney Islands" : LocalAuthority("Orkney Islands",22190,988.8,22)}



#if the following file doesn't exist get it, it contains the school data
data_file = Path("./00548378.xlsx")
try:
    data_file.resolve(strict=True)
except FileNotFoundError:
    url = 'https://www2.gov.scot/Resource/0054/00548378.xlsx'
    urlretrieve(url, data_file)
    #wget https://www2.gov.scot/Topics/Statistics/Browse/School-Education/schoolestatestats/00548378.xlsx
    print ("File doesn't exist")
else:
    print ("File exists")

workbook = openpyxl.load_workbook("00548378.xlsx")

# the names of the local authority in this spreadsheet doesn't match the names we have
authority_name_map = { "Aberdeen City" : "City of Aberdeen",
"Argyll & Bute" : "Argyll and Bute",
"Dumfries & Galloway" : "Dumfries and Galloway",
"Dundee City" : "City of Dundee",
"East Lothian(2)" : "East Lothian",
"Edinburgh, City of":"City of Edinburgh",
"Na h-Eileanan Siar" : "Western Isles (Na h-Eileanan Siar)",
"Glasgow City" : "City of Glasgow",
"Perth & Kinross" : "Perth and Kinross"}

# Get the number of schools in each local authority from Table 1
num_school_worksheet = workbook["Table 1"]
for row in num_school_worksheet.iter_rows(min_row=4, max_col=8, max_row=40):

    if row[0].value is not None:

        authority_name = row[0].value
        if authority_name in authority_name_map:
            authority_name = authority_name_map[row[0].value]

        localAuthorities[authority_name].primary_schools = int(row[1].value)
        localAuthorities[authority_name].secondary_schools = int(row[2].value)
        localAuthorities[authority_name].total_schools = int(row[4].value)

# Get the population of school pupils from Table 8
school_pop_worksheet = workbook["Table 8"]
for row in school_pop_worksheet.iter_rows(min_row=5, max_col=17, max_row=2496):

    authority_name = row[2].value
    if authority_name in authority_name_map:
        authority_name = authority_name_map[row[2].value]

    num_pupils = 0
    try:
        num_pupils = int(row[14].value)
    except ValueError:
        pass # there is a # in the Capacity percent column
    if num_pupils > 0:
        if "Primary" == row[3].value:
            localAuthorities[authority_name].primary_school_total += num_pupils
            localAuthorities[authority_name].primary_school_pupils.append(num_pupils)
        elif "Secondary" == row[3].value:
            localAuthorities[authority_name].secondary_school_total += num_pupils
            localAuthorities[authority_name].secondary_school_pupils.append(num_pupils)
        elif "Special" == row[3].value:
            localAuthorities[authority_name].special_school_total += num_pupils
            localAuthorities[authority_name].special_school_pupils.append(num_pupils)


# The algorithm is now:
# Assuming that an average class size is 30 pupils and that there is complete mixing within a class and 60% mixing
# between classes in a school the mixing (ave number of contacts a pupil will have in a school) is
# 30*1.0 + (school_size-30)*0.6
# and multiplying this by the number of pupils in the school will give the number of contacts within the school.
#
# The number of contacts between schools in a local authority, assuming that the contact rate is 5%, is
# 0.5 * 0.05 * (number of pupils in local authority - number pupils in the school)
# where the factor of 0.5 ensures we do not double count between school contacts.
#
# The number of contacts is presented as the average number of contacts a school age child will have as a 
# fraction (percentage) of the size of the school.
#
# The total number of contacts in a local authority is an average of all the contacts wєighted according to the 
# school size.
# 
# A similar weighting across local authorities in a health board is done.

# The weighting factors are asssumed to be:
within_class_mixing = 1.0
within_school_mixing = 0.4
within_la_mixing = 0.01
ave_class_size = 30

mixing = 0.0
total_pupils = 0
print(f"Mixing rates per health board")
for hb, laList in healthBoardsAndLocalAuthoriesMap.items():

    total_pupils_in_hb = 0
    contacts_in_hb = 0

    for local_authority_name in laList:
        la = localAuthorities[local_authority_name]
        school_sizes = la.primary_school_pupils + la.secondary_school_pupils + la.special_school_pupils

        total_pupils_in_la = sum(school_sizes)
        total_pupils_in_hb += total_pupils_in_la
        school_age_contacts = 0.0

        for num_pupils in school_sizes:
            within_school = (within_class_mixing*ave_class_size) + within_school_mixing*(num_pupils-ave_class_size)
            between_schools = 0.5*within_la_mixing*(total_pupils_in_la-num_pupils)
            school_age_contacts +=  ((num_pupils/total_pupils_in_la) * (within_school + between_schools))/num_pupils
            # school_age_contacts is the average numebr of contacts a single pupil will have as a percentage of the school size

        contacts_in_hb += school_age_contacts*(total_pupils_in_la/total_pupils_in_hb)

    print (f"{hb}, {contacts_in_hb:.2f}")
    mixing += contacts_in_hb*total_pupils_in_hb
    total_pupils += total_pupils_in_hb

print(f"\n[5,18) -> [5,18) mixing = {(mixing/total_pupils):.2f}")


"""
# we can get the number of primary (0-11year olds) and secondary (12-17 yo) in each health board.
print (f"#Health Board, Num Primary school pupils, Num Secondary School pupils, Total (includes special schools)")
for hb, laList in healthBoardsAndLocalAuthoriesMap.items():
    num_prim_pupils = 0
    num_sec_pupils = 0
    num_pupils = 0
    for la in laList:
        num_prim_pupils += localAuthorities[la].primary_school_total
        num_sec_pupils += localAuthorities[la].secondary_school_total
        num_pupils += localAuthorities[la].special_school_total

    print (f"{hb}, {num_prim_pupils:.2f}, {num_sec_pupils:.2f}, {num_pupils:.2f}")
"""

